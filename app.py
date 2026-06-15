import streamlit as st
import google.generativeai as genai
import openpyxl
import io
import json

# 1. Configuración de la interfaz visual de la app
st.set_page_config(page_title="Asistente de Laboratorio IQ", page_icon="🧪", layout="wide")
st.title("🧪 Calificador Avanzado de Laboratorio (Versión SAEw Profesional)")
st.write("Configura tus criterios con documentos guía, sube los informes de tus alumnos y automatiza el acta del SAEw.")

# Cargar la llave de la IA de forma segura
try:
    GENAI_KEY = st.secrets["CONEXION_IA"]
except Exception:
    st.error("Falta configurar la llave secreta 'CONEXION_IA' en Streamlit Cloud.")
    st.stop()

# Configurar la IA de Google
genai.configure(api_key=GENAI_KEY)

# Estructura de pestañas para organizar la interfaz de Carla
tab1, tab2 = st.tabs(["📋 Proceso de Calificación", "⚙️ Guía de Uso"])

with tab1:
    # 2. ESTACIÓN DE CRITERIOS MULTIMEDIAL (Paso 1)
    st.header("Paso 1: Define el Criterio y Formato de Evaluación")
    col_crit1, col_crit2 = st.columns([1, 1])
    
    with col_crit1:
        modelo_respuestas = st.text_area(
            "✍️ Indicaciones en texto o penalizaciones específicas:",
            placeholder="Ejemplo: Restar 1 punto si no incluyen nomenclatura de variables. Respuesta correcta de la difusión: 0.090 cm²/s.",
            height=150
        )
    
    with col_crit2:
        archivos_criterio = st.file_uploader(
            "📁 Sube los archivos guía (Soluciones a mano, PDF modelo de 10/10, o formato base oficial):",
            type=["png", "jpg", "jpeg", "pdf", "docx"],
            accept_multiple_files=True,
            key="criterios_maestros"
        )
        if archivos_criterio:
            st.success(f"📎 {len(archivos_criterio)} archivo(s) de referencia cargado(s) con éxito.")

    st.markdown("---")

    # 3. CARGA DE INFORMES DE ESTUDIANTES (Paso 2)
    st.header("Paso 2: Carga de Documentos de Estudiantes")
    archivos_subidos = st.file_uploader(
        "Arrastra aquí los informes grupales en PDF o fotos de coloquios de los alumnos:", 
        type=["png", "jpg", "jpeg", "pdf"], 
        accept_multiple_files=True,
        key="informes_alumnos"
    )

    st.markdown("---")

    # 4. PARSING ESTRICTO DEL EXCEL DEL SAEW (Paso 3)
    st.header("Paso 3: Acta de Calificaciones del SAEw")
    excel_file = st.file_uploader("Arrastra el archivo Excel original exportado directamente del SAEw:", type=["xlsx"])

    practica_seleccionada = None
    header_row_idx = None
    practice_row_idx = None

    if excel_file:
        try:
            bytes_in_mem = excel_file.read()
            wb_reader = openpyxl.load_workbook(io.BytesIO(bytes_in_mem))
            sheet_reader = wb_reader.active
            
            # BÚSQUEDA ESTRICTA: Evita confundirse con textos institucionales largos como los de la fila 5
            for r in range(1, 15):
                row_vals = [str(sheet_reader.cell(row=r, column=c).value).strip().lower() for c in range(1, 25)]
                if "estudiante" in row_vals or "código" in row_vals or "nro." in row_vals:
                    header_row_idx = r
                    break
            
            if header_row_idx and header_row_idx > 1:
                practice_row_idx = header_row_idx - 1
                listado_practicas = []
                for c in range(1, sheet_reader.max_column + 1):
                    val = sheet_reader.cell(row=practice_row_idx, column=c).value
                    if val and str(val).strip() and str(val).strip() not in listado_practicas:
                        if "escuela" not in str(val).lower() and "fuente" not in str(val).lower():
                            listado_practicas.append(str(val).strip())
                
                if listado_practicas:
                    practica_seleccionada = st.selectbox("🎯 ¿Qué práctica del SAEw vas a evaluar hoy?", listado_practicas)
                else:
                    st.error("No se detectaron nombres de prácticas válidos en la fila superior a la cabecera.")
            else:
                st.error("No se reconoció la estructura del SAEw. Asegúrate de que el archivo contenga la columna 'Estudiante'.")
                
            excel_file.seek(0) # Resetear puntero
        except Exception as e:
            st.error(f"Error analizando la estructura del archivo Excel: {e}")

    st.markdown("---")

    # 5. EJECUCIÓN DEL MOTOR DE INTELIGENCIA ARTIFICIAL MULTIMODAL
    if archivos_subidos and excel_file and practica_seleccionada and (modelo_respuestas or archivos_criterio):
        if st.button("🚀 Iniciar Calificación Inteligente"):
            progreso = st.progress(0)
            status_text = st.empty()
            
            # Preparar los archivos de criterio (Rúbricas/Modelos) en memoria una sola vez
            payload_criterios = []
            if archivos_criterio:
                for arch_ref in archivos_criterio:
                    ref_bytes = arch_ref.read()
                    payload_criterios.append({"mime_type": arch_ref.type if arch_ref.type else "application/pdf", "data": ref_bytes})
                    arch_ref.seek(0)
            
            # Cargar el libro de trabajo para edición
            wb = openpyxl.load_workbook(io.BytesIO(excel_file.read()))
            sheet = wb.active
            
            # Encontrar columna de nombres de estudiantes
            col_nombre = 4
            for c in range(1, 25):
                val_h = str(sheet.cell(row=header_row_idx, column=c).value).lower()
                if "estudiante" in val_h:
                    col_nombre = c
                    break
            
            # Encontrar columna de la sección de la práctica seleccionada
            start_col = None
            for c in range(1, sheet.max_column + 1):
                val_p = sheet.cell(row=practice_row_idx, column=c).value
                if val_p and str(val_p).strip().lower() == practica_seleccionada.lower():
                    start_col = c
                    break
            
            # Localizar subcolumna 'I' (Informe) dentro de la práctica elegida
            col_nota = None
            if start_col:
                for c in range(start_col, start_col + 5):
                    val_sub = str(sheet.cell(row=header_row_idx, column=c).value).lower().strip()
                    if val_sub == 'i' or val_sub == 'i2' or 'inf' in val_sub:
                        col_nota = c
                        break
            
            if not col_nota:
                col_nota = start_col + 2 if start_col else 10 # Fallback posicional seguro
            
            # Crear columna de observaciones al final de la hoja para conservar el formato limpio del SAEw
            col_feedback = sheet.max_column + 1
            sheet.cell(row=header_row_idx, column=col_feedback).value = f"Feedback - {practica_seleccionada}"
            
            total_archivos = len(archivos_subidos)
            
            # Procesar cada informe individualmente
            for index, archivo_alumno in enumerate(archivos_subidos):
                status_text.text(f"Analizando trabajo {index+1} de {total_archivos}: {archivo_alumno.name}")
                
                alumno_bytes = archivo_alumno.read()
                model = genai.GenerativeModel('gemini-1.5-flash')
                
                # Diseño de prompt multimodal avanzado
                prompt = f"""
                Actúas como un docente de laboratorio altamente detallista de Ingeniería Química de la Escuela Politécnica Nacional (EPN).
                Tu misión es calificar el 'INFORME DEL ESTUDIANTE' que se encuentra adjunto al final de los elementos visuales.
                
                Para calificar de manera justa y exacta, básate estrictamente en los siguientes insumos de referencia que te ha provisto el docente:
                1. Reglas e instrucciones en texto: {modelo_respuestas}
                2. Los primeros archivos adjuntos enviados en esta solicitud corresponden a guías oficiales de formato, soluciones ideales de problemas o informes calificados con nota perfecta (10/10) que sirven como tu patrón de comparación.
                
                Compara minuciosamente el formato, estructura, resultados y discusiones técnicas del 'INFORME DEL ESTUDIANTE' frente a dichos patrones de referencia.
                
                Busca en las primeras páginas del 'INFORME DEL ESTUDIANTE' los nombres y apellidos de todos los integrantes del grupo de laboratorio que firman el trabajo.
                
                Devuelve la respuesta EXCLUSIVAMENTE en este formato JSON estructurado, sin bloques de código ```json o texto de introducción:
                {{
                  "nombres": ["Nombre Apellido Alumno 1", "Nombre Apellido Alumno 2"],
                  "nota": 8.7,
                  "justificacion": "Explicación concisa y de alto valor técnico sobre los aciertos o errores cometidos comparados con la solución y formato de referencia."
                }}
                """
                
                # Construcción dinámica del paquete de datos para Gemini
                paquete_ia = []
                paquete_ia.extend(payload_criterios) # Primero inyectamos las guías máster de Carla
                paquete_ia.append({"mime_type": archivo_alumno.type, "data": alumno_bytes}) # Al final el archivo del alumno
                paquete_ia.append(prompt)
                
                try:
                    response = model.generate_content(paquete_ia)
                    texto_limpio = response.text.replace("```json", "").replace("```", "").strip()
                    datos = json.loads(texto_limpio)
                    
                    lista_nombres = datos.get("nombres", [])
                    nota_final = datos.get("nota", 0)
                    feedback_final = datos.get("justificacion", "")
                    
                    # Cruzar nombres detectados con las filas del Excel del SAEw
                    for nombre_detectado in lista_nombres:
                        for row in range(header_row_idx + 1, sheet.max_row + 1):
                            celda_nombre = str(sheet.cell(row=row, column=col_nombre).value).lower().strip()
                            # Validación cruzada inteligente de nombres
                            if nombre_detectado.lower().strip() in celda_nombre or celda_nombre in nombre_detectado.lower().strip():
                                sheet.cell(row=row, column=col_nota).value = float(nota_final)
                                sheet.cell(row=row, column=col_feedback).value = feedback_final
                                break
                                
                except Exception as e:
                    st.warning(f"Aviso en {archivo_alumno.name}: No se pudo procesar automáticamente. Verifica el formato del documento. Detalles: {str(e)}")
                
                progreso.progress((index + 1) / total_archivos)
            
            status_text.text("¡Todo el paralelo ha sido calificado con éxito!")
            st.success(f"Las notas e informes fueron vinculados correctamente a la columna del examen/informe de: {practica_seleccionada}")
            
            # Preparar descarga del archivo final procesado
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            st.download_button(
                label="📥 Descargar Acta de Notas SAEw Actualizada",
                data=output,
                file_name=f"Acta_Procesada_{excel_file.name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

with tab2:
    st.subheader("💡 ¿Cómo funciona la Estación de Criterios Multimedial?")
    st.write("""
    Ahora puedes entrenar a la IA antes de cada calificación de tres maneras simultáneas:
    1. **Formato Base:** Sube el documento vacío de la rúbrica institucional. La IA analizará la estructura del informe del estudiante para ver si respeta las secciones requeridas.
    2. **Soluciones Gráficas:** Si resolviste los cálculos de transferencia de calor en un papel o pizarra, tómale una foto y súbela en el Paso 1. La IA la usará para comprobar las respuestas numéricas.
    3. **Informes de Muestra:** Puedes subir un informe excelente de un semestre pasado para guiar el nivel de exigencia en la sección de 'Resultados y Discusión'.
    """)
